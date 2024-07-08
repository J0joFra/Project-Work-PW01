import requests
from openpyxl import Workbook

def fetch_data(url):
    response = requests.post(url)
    if response.status_code == 200:
        data = response.json()
        return [
            (
                item.get('pcpc_ingredientid', ''),
                item.get('pcpc_ingredientname', ''),
                f"https://cir-reports.cir-safety.org/cir-ingredient-status-report/?id={item.get('pcpc_ingredientid', '')}"
            )
            for item in data.get('results', [])
        ]
    else:
        print(f"Errore nella richiesta: {response.status_code} - {response.reason}")
        return []

def create_workbook():
    wb = Workbook()
    ws = wb.active
    ws.title = 'CIR Ingredients Report'
    headers = ['pcpc_ingredientid', 'pcpc_ingredientname', 'link']  # Intestazioni colonne

    # Scrittura intestazioni nel foglio di lavoro
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    print("Intestazione scritta")
    return wb, ws

def write_data_to_excel(ws, data, start_row):
    row_idx = start_row
    for element in data:
        for col_idx, value in enumerate(element, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
        row_idx += 1
        if row_idx % 1000 == 0:
            print(f"Scrittura di {row_idx} righe")
    return row_idx

def remove_duplicates(ws):
    seen = set()
    duplicates = []
    for row in range(2, ws.max_row + 1):
        pcpc_ingredientid = ws.cell(row=row, column=1).value
        if pcpc_ingredientid in seen:
            duplicates.append(row)
        else:
            seen.add(pcpc_ingredientid)

    for row in reversed(duplicates):
        ws.delete_rows(row) 
    print(f"Rimossi {len(duplicates)} duplicati")

def main():
    # URL delle richieste POST per ottenere i dati JSON
    urls = [
        "https://cir-reports.cir-safety.org/FetchCIRReports/", #a-p
        "https://cir-reports.cir-safety.org/FetchCIRReports/?&pagingcookie=%26lt%3bcookie%20page%3d%26quot%3b1%26quot%3b%26gt%3b%26lt%3bpcpc_name%20last%3d%26quot%3bPEG-50%20Stearate%26quot%3b%20first%3d%26quot%3b1%2c10-Decanediol%26quot%3b%20%2f%26gt%3b%26lt%3bpcpc_ingredientidname%20last%3d%26quot%3bPEG-50%20Stearate%26quot%3b%20first%3d%26quot%3b1%2c10-Decanediol%26quot%3b%20%2f%26gt%3b%26lt%3bpcpc_cirrelatedingredientsid%20last%3d%26quot%3b%7bC223037E-F278-416D-A287-2007B9671D0C%7d%26quot%3b%20first%3d%26quot%3b%7b940AF697-52B5-4A3A-90A6-B9DB30EF4A7E%7d%26quot%3b%20%2f%26gt%3b%26lt%3b%2fcookie%26gt%3b&page=2"
    ]

    try:
        # Crea un nuovo workbook Excel e scrivi l'intestazione
        wb, ws = create_workbook()

        # Estrai e scrivi i dati dai vari URL
        row_idx = 2
        for url in urls:
            desired_elements = fetch_data(url)
            row_idx = write_data_to_excel(ws, desired_elements, start_row=row_idx)

        # Rimuovi i duplicati
        remove_duplicates(ws)

        # Salva il workbook Excel
        filename = 'CIR_Ingredients_Report.xlsx'
        wb.save(filename=filename)
        print("File Excel creato con successo!")

    except Exception as e:
        print(f"Si è verificato un errore durante l'esecuzione: {str(e)}")

# Esegui la funzione principale
if __name__ == "__main__":
    main()




